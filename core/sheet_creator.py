"""Sheet creation logic for attendance books"""

import calendar
from datetime import datetime
import holidays
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from utils.constants import (
    MONTH_NAMES_ARABIC, DAY_NAMES_ARABIC, DAILY_HEADERS,
    DAILY_COLUMN_WIDTHS, MONTHLY_COLUMN_WIDTHS, FONT_SIZES
)


class SheetCreator:
    """Handles creation of daily and monthly attendance sheets"""
    
    @staticmethod
    def create_daily_sheet(wb, sheet_name, names, date):
        """Create a daily attendance sheet"""
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.rightToLeft = True
        
        # Setup page header
        header_text = f"دفتر حضور وغياب الطلاب - {sheet_name} - التاريخ: {date}"
        ws.oddHeader.center.text = header_text
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.size = 14
        
        # Formatting
        bold_font = Font(bold=True, size=FONT_SIZES['header'])
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Column headers
        for i, header in enumerate(DAILY_HEADERS, start=1):
            cell = ws.cell(row=1, column=i)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border
        
        # Insert names and data
        for i, name in enumerate(names, start=2):
            ws.cell(row=i, column=1, value=i-1).border = border
            ws.cell(row=i, column=1).alignment = center_align
            ws.cell(row=i, column=2, value=name).border = border
            
            for col in range(3, 7):
                ws.cell(row=i, column=col).border = border
                if col in [3, 4]:
                    ws.cell(row=i, column=col).alignment = center_align
        
        # Set column widths
        for col, width in DAILY_COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width
        
        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        return ws
    
    @staticmethod
    def create_monthly_sheet(wb, sheet_name, names, year, month, highlight_holidays=True):
        """Create a monthly attendance sheet"""
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.rightToLeft = True
        
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Setup page header
        header_text = f"دفتر الحضور الشهري - {sheet_name} - {MONTH_NAMES_ARABIC[month-1]} {year}"
        ws.oddHeader.center.text = header_text
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.size = 14
        
        # Get Jordan holidays
        jordan_holidays = holidays.Jordan(years=year) if highlight_holidays else {}
        
        # Formatting
        bold_font = Font(bold=True, size=FONT_SIZES['header'])
        small_font = Font(bold=True, size=FONT_SIZES['day_name'])
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        holiday_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Number column header
        ws.cell(row=2, column=1, value='الرقم').font = bold_font
        ws.cell(row=2, column=1).alignment = center_align
        ws.cell(row=2, column=1).border = border
        ws.cell(row=2, column=1).fill = normal_fill
        
        # Name column header
        ws.cell(row=2, column=2, value='الاسم').font = bold_font
        ws.cell(row=2, column=2).alignment = center_align
        ws.cell(row=2, column=2).border = border
        ws.cell(row=2, column=2).fill = normal_fill
        
        # Day headers
        for day in range(1, days_in_month + 1):
            col = day + 2
            date_obj = datetime(year, month, day)
            day_of_week = date_obj.weekday()
            arabic_day_index = (day_of_week + 1) % 7
            day_name = DAY_NAMES_ARABIC[arabic_day_index]
            
            is_friday = arabic_day_index == 5
            is_saturday = arabic_day_index == 6
            is_holiday = date_obj.date() in jordan_holidays
            is_weekend_or_holiday = is_friday or is_saturday or is_holiday
            
            cell_fill = holiday_fill if (highlight_holidays and is_weekend_or_holiday) else normal_fill
            
            day_name_cell = ws.cell(row=1, column=col, value=day_name)
            day_name_cell.font = small_font
            day_name_cell.alignment = center_align
            day_name_cell.border = border
            day_name_cell.fill = cell_fill
            
            day_num_cell = ws.cell(row=2, column=col, value=str(day))
            day_num_cell.font = bold_font
            day_num_cell.alignment = center_align
            day_num_cell.border = border
            day_num_cell.fill = cell_fill
        
        # Insert names
        for i, name in enumerate(names, start=3):
            num_cell = ws.cell(row=i, column=1, value=i-2)
            num_cell.border = border
            num_cell.alignment = center_align
            num_cell.fill = normal_fill
            
            name_cell = ws.cell(row=i, column=2, value=name)
            name_cell.font = Font(size=FONT_SIZES['name'])
            name_cell.border = border
            name_cell.alignment = right_align
            name_cell.fill = normal_fill
            
            for day in range(1, days_in_month + 1):
                col = day + 2
                date_obj = datetime(year, month, day)
                day_of_week = date_obj.weekday()
                arabic_day_index = (day_of_week + 1) % 7
                is_friday = arabic_day_index == 5
                is_saturday = arabic_day_index == 6
                is_holiday = date_obj.date() in jordan_holidays
                is_weekend_or_holiday = is_friday or is_saturday or is_holiday
                
                cell_fill = holiday_fill if (highlight_holidays and is_weekend_or_holiday) else normal_fill
                
                cell = ws.cell(row=i, column=col)
                cell.border = border
                cell.alignment = center_align
                cell.fill = cell_fill
        
        # Add total column
        total_col = days_in_month + 3
        
        total_header = ws.cell(row=1, column=total_col, value="المجموع")
        total_header.font = bold_font
        total_header.alignment = center_align
        total_header.border = border
        total_header.fill = normal_fill
        
        total_second_row = ws.cell(row=2, column=total_col, value="Total")
        total_second_row.font = bold_font
        total_second_row.alignment = center_align
        total_second_row.border = border
        total_second_row.fill = normal_fill
        
        for i in range(3, len(names) + 3):
            total_cell = ws.cell(row=i, column=total_col)
            total_cell.border = border
            total_cell.alignment = center_align
            total_cell.fill = normal_fill
            total_cell.font = Font(size=FONT_SIZES['name'])
        
        # Set column widths
        ws.column_dimensions['A'].width = MONTHLY_COLUMN_WIDTHS['number']
        ws.column_dimensions['B'].width = MONTHLY_COLUMN_WIDTHS['name']
        
        for day in range(1, days_in_month + 1):
            col_letter = get_column_letter(day + 2)
            ws.column_dimensions[col_letter].width = MONTHLY_COLUMN_WIDTHS['day']
        
        total_col_letter = get_column_letter(total_col)
        ws.column_dimensions[total_col_letter].width = MONTHLY_COLUMN_WIDTHS['total']
        
        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = "1:2"
        
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
        return ws
