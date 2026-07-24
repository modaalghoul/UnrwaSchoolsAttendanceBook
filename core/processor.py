"""Attendance processing logic"""

from PyQt6.QtCore import QThread, pyqtSignal
import xlrd
from openpyxl import Workbook

from core.sheet_creator import SheetCreator
from utils.constants import MAX_NAMES


class AttendanceProcessor(QThread):
    """Background thread for processing attendance data"""
    
    progress_updated = pyqtSignal(int)
    status_updated = pyqtSignal(str)
    finished = pyqtSignal(str)
    error_occurred = pyqtSignal(str)

    def __init__(self, input_file, output_file, report_type, date_info, highlight_holidays=True):
        super().__init__()
        self.input_file = input_file
        self.output_file = output_file
        self.report_type = report_type
        self.date_info = date_info
        self.highlight_holidays = highlight_holidays
        
    def extract_names_from_sheet(self, sheet):
        """Extract names from worksheet"""
        names = []
        name_column = None
        start_row = None
        
        # Search for "الاسم" cell
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                try:
                    cell_value = str(sheet.cell(row, col).value).strip()
                    if cell_value == "الاسم":
                        name_column = col
                        start_row = row + 1
                        break
                except:
                    continue
            if name_column is not None:
                break
        
        if name_column is None:
            return names
            
        # Extract names
        for row in range(start_row, sheet.nrows):
            try:
                name = str(sheet.cell(row, name_column).value).strip()
                if name and name != "" and name.lower() != "none":
                    names.append(name)
                else:
                    break
            except:
                break
                
        return names[:MAX_NAMES]
    
    def run(self):
        """Main processing logic"""
        try:
            self.status_updated.emit("بدء معالجة الملف...")
            self.progress_updated.emit(10)
            
            # Read input Excel file
            workbook = xlrd.open_workbook(self.input_file)
            sheets_data = {}
            
            # Extract names from each sheet
            total_sheets = len(workbook.sheet_names())
            for i, sheet_name in enumerate(workbook.sheet_names()):
                self.status_updated.emit(f"معالجة الشعبة: {sheet_name}")
                sheet = workbook.sheet_by_name(sheet_name)
                names = self.extract_names_from_sheet(sheet)
                if names:
                    sheets_data[sheet_name] = names
                
                progress = 10 + (i + 1) * 40 // total_sheets
                self.progress_updated.emit(progress)
            
            if not sheets_data:
                raise Exception("لم يتم العثور على أي أسماء في الملف")
            
            self.status_updated.emit("إنشاء ملف دفتر الغياب...")
            
            # Create new Excel file
            wb = Workbook()
            if wb.worksheets:
                wb.remove(wb.active)
            
            # Create attendance sheets
            sheet_count = 0
            total_operations = len(sheets_data)
            
            for sheet_name, names in sheets_data.items():
                if self.report_type == "daily":
                    date_str = self.date_info.toString("yyyy-MM-dd")
                    SheetCreator.create_daily_sheet(wb, sheet_name, names, date_str)
                else:
                    year = self.date_info['year']
                    month = self.date_info['month']
                    SheetCreator.create_monthly_sheet(
                        wb, sheet_name, names, year, month, self.highlight_holidays
                    )
                
                sheet_count += 1
                progress = 50 + (sheet_count * 40 // total_operations)
                self.progress_updated.emit(progress)
            
            # Save file
            self.status_updated.emit("حفظ الملف...")
            wb.save(self.output_file)
            
            self.progress_updated.emit(100)
            self.finished.emit(f"تم إنشاء دفتر الغياب بنجاح!\nتم إنشاء {len(sheets_data)} ورقة حضور")
            
        except Exception as e:
            self.error_occurred.emit(f"خطأ في المعالجة: {str(e)}")
